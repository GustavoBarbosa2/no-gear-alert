from ultralytics import YOLO
import cv2
import os
import time
from datetime import datetime
from openpyxl import Workbook, load_workbook

EXCEL_PATH = "registo_alertas.xlsx"

EPIS_OBRIGATORIOS = ["helmet", "vest", "glasses", "mask"]  

def init_excel():
    if not os.path.exists(EXCEL_PATH):
        wb = Workbook()
        ws = wb.active
        ws.append(["Data/Hora", "EPIs Detetados", "EPIs em Falta", "Alerta"])
        wb.save(EXCEL_PATH)

def registar_alerta(epis_detectados):
    agora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    epis_detectados = [epi.lower() for epi in epis_detectados]
    faltam = [epi for epi in EPIS_OBRIGATORIOS if epi not in epis_detectados]
    alerta_msg = "Todos os EPIs presentes" if not faltam else f"Faltam: {', '.join(faltam)}"

    wb = load_workbook(EXCEL_PATH)
    ws = wb.active
    ws.append([
        agora,
        ", ".join(epis_detectados) if epis_detectados else "Nenhum",
        ", ".join(faltam) if faltam else "Nenhum",
        alerta_msg
    ])
    wb.save(EXCEL_PATH)
    print(f"Alerta registado: {alerta_msg}")

def main():
    init_excel()
    model = YOLO("runs/detect/train/weights/best.pt")
    cap = cv2.VideoCapture(0)

    if not cap.isOpened():
        print("Não foi possível abrir a webcam.")
        return

    print("Webcam ligada. Pressiona 'q' para sair.")

    last_alert_time = 0
    alert_interval = 5  

    while True:
        ret, frame = cap.read()
        if not ret:
            break

        results = model(frame)
        annotated_frame = results[0].plot()
        boxes = results[0].boxes
        classes = results[0].names

        epis_detectados = []
        if boxes.cls is not None:
            for cls_id in boxes.cls:
                cls_name = classes[int(cls_id)]
                if cls_name not in epis_detectados:
                    epis_detectados.append(cls_name)

        faltam = [epi for epi in EPIS_OBRIGATORIOS if epi not in epis_detectados]
        if faltam and (time.time() - last_alert_time > alert_interval):
            registar_alerta(epis_detectados)
            last_alert_time = time.time()

        cv2.imshow("Detecção de EPIs", annotated_frame)

        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    cap.release()
    cv2.destroyAllWindows()

if __name__ == "__main__":
    main()
