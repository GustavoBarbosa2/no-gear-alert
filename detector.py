from ultralytics import YOLO
import cv2
from datetime import datetime
from openpyxl import Workbook, load_workbook
import os

EXCEL_PATH = "registo_alertas.xlsx"

EPIS_OBRIGATORIOS = ["helmet", "vest", "glasses", "mask"]

def init_excel():
    if not os.path.exists(EXCEL_PATH):
        wb = Workbook()
        ws = wb.active
        ws.append(["Data/Hora", "EPIs Detetados", "EPIs em Falta", "Alerta"])
        wb.save(EXCEL_PATH)

def registar_alerta(epis_detectados):
    agora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    epis_detectados = [epi.lower() for epi in epis_detectados]
    faltam = [epi for epi in EPIS_OBRIGATORIOS if epi not in epis_detectados]

    alerta_msg = "Todos os EPIs presentes" if not faltam else f"Faltam: {', '.join(faltam)}"

    wb = load_workbook(EXCEL_PATH)
    ws = wb.active
    ws.append([
        agora,
        ", ".join(epis_detectados) if epis_detectados else "Nenhum",
        ", ".join(faltam) if faltam else "Nenhum",
        alerta_msg
    ])
    wb.save(EXCEL_PATH)

    if faltam:
        print(f"Alerta registado: {alerta_msg}")
    else:
        print(f"Tudo OK: {alerta_msg}")


def main():
    init_excel()
    model = YOLO('runs/detect/train/weights/best.pt')

    video_path = 'video2.mp4'
    cap = cv2.VideoCapture(video_path)

    if not cap.isOpened():
        print(f"Erro ao abrir o vídeo: {video_path}")
        return

    last_alert_time = 0
    alert_interval = 5  

    while True:
        ret, frame = cap.read()
        if not ret:
            break

        results = model(frame)
        annotated_frame = results[0].plot()

        boxes = results[0].boxes
        names = results[0].names
        epis_detectados = []

        if boxes.cls is not None:
            for cls_id in boxes.cls:
                cls_name = names[int(cls_id)]
                if cls_name not in epis_detectados:
                    epis_detectados.append(cls_name)

        faltam = [epi for epi in EPIS_OBRIGATORIOS if epi not in epis_detectados]
        if faltam:
            current_time = cv2.getTickCount() / cv2.getTickFrequency()
            if current_time - last_alert_time > alert_interval:
                registar_alerta(epis_detectados)
                last_alert_time = current_time

        cv2.imshow('Detecção de EPI (vídeo)', annotated_frame)

        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    cap.release()
    cv2.destroyAllWindows()

if __name__ == '__main__':
    main()
